from multiprocessing import Value
from urllib import request
from django.http import HttpResponse
from django.contrib.auth.hashers import make_password
from django.contrib.auth import update_session_auth_hash
from django.shortcuts import render,redirect, get_object_or_404
from django.contrib import messages
from django.template import loader
from ..models import Commission, LeadUploadExcel, SourceMaster,Users, DocumentUpload, Branch, Leads, QuotationCustomer
from empPortal.model import BankDetails
from ..forms import DocumentUploadForm
from django.core.mail import send_mail
from django.core.mail import EmailMessage
from django.utils.timezone import now
from django.contrib.auth import authenticate, login ,logout
from django.core.files.storage import FileSystemStorage
import re
import requests
from fastapi import FastAPI, File, UploadFile
import fitz
import openai
import time
import json
from django.http import JsonResponse
import os
import zipfile
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.core.cache import cache
from django.db import connection
import logging
logger = logging.getLogger(__name__)
import os
import pdfkit
from django.template.loader import render_to_string
from pprint import pprint 
from django.core.paginator import Paginator
from django.db.models import Q
from empPortal.model import Referral

import pandas as pd
from django.core.files.storage import default_storage
import openpyxl
from django.db.models import Max
import re,logging
from dateutil import parser
logger = logging.getLogger(__name__)
OPENAI_API_KEY = settings.OPENAI_API_KEY
from django_q.tasks import async_task
from ..models import Users, LeadUploadExcel
from datetime import datetime, timedelta
from django.db.models import F, Value, CharField
from django.db.models.functions import Concat, Coalesce
from ..models import PolicyInfo
app = FastAPI()

def dictfetchall(cursor):
    "Returns all rows from a cursor as a dict"
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def index(request):
    if not request.user.is_authenticated:
        return redirect('login')

    per_page = request.GET.get('per_page', 10)
    search_field = request.GET.get('search_field', '')
    search_query = request.GET.get('search_query', '')
    global_search = request.GET.get('global_search', '').strip()
    shorting = request.GET.get('shorting', '')  # Get sorting preference

    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10

    # Base queryset
    if request.user.role_id != 1:
        leads = Leads.objects.filter(created_by=request.user.id)
    else:
        leads = Leads.objects.all()

    # Global search
    if global_search:
        leads = leads.filter(
            Q(name_as_per_pan__icontains=global_search) |
            Q(email_address__icontains=global_search) |
            Q(mobile_number__icontains=global_search) |
            Q(pan_card_number__icontains=global_search) |
            Q(state__icontains=global_search) |
            Q(city__icontains=global_search) |
            Q(lead_id__icontains=global_search)
        )

    # Field-specific search
    if search_field and search_query:
        filter_args = {f"{search_field}__icontains": search_query}
        leads = leads.filter(**filter_args)

     # Apply filters based on form input
    """if 'lead_id' in request.GET and request.GET['lead_id']:
        leads = leads.filter(lead_id__icontains=request.GET['lead_id'])
    if 'name_as_per_pan' in request.GET and request.GET['name_as_per_pan']:
        leads = leads.filter(name_as_per_pan__icontains=request.GET['name_as_per_pan'])    
    if 'pan_card_number' in request.GET and request.GET['pan_card_number']:
        leads = leads.filter(pan_card_number__icontains=request.GET['pan_card_number'])
    if 'email_address' in request.GET and request.GET['email_address']:
        leads = leads.filter(email_address__icontains=request.GET['email_address'])
    if 'mobile_number' in request.GET and request.GET['mobile_number']:
        leads = leads.filter(mobile_number__icontains=request.GET['mobile_number']) """

    # Get filter inputs
    lead_id = request.GET.get('lead_id', '')
    name = request.GET.get('name_as_per_pan', '')
    pan = request.GET.get('pan_card_number', '')
    email = request.GET.get('email_address', '')
    mobile = request.GET.get('mobile_number', '')
    sales_manager = request.GET.get('sales_manager', '')
    agent_name = request.GET.get('agent_name', '')
    policy_number = request.GET.get('policy_number', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    insurance_company = request.GET.get('insurance_company', '')
    policy_type = request.GET.get('policy_type', '')
    vehicle_type = request.GET.get('vehicle_type', '')
    upcoming_renewals = request.GET.get('upcoming_renewals', '')
    lead_type = request.GET.get('lead_type')
    motor_type = request.GET.get('motor_type')

    #today = datetime.today().date()
    #after_30_days = today + timedelta(days=30)
    # Apply filters
    if lead_id:
        leads = leads.filter(lead_id__icontains=lead_id)
    if name:
        leads = leads.filter(name_as_per_pan__icontains=name)
    if pan:
        leads = leads.filter(pan_card_number__icontains=pan)
    if email:
        leads = leads.filter(email_address__icontains=email)
    if mobile:
        leads = leads.filter(mobile_number__icontains=mobile)
    if sales_manager:
        leads = leads.filter(sales_manager__user_name=sales_manager)
    if agent_name:
        leads = leads.filter(agent_name=agent_name)
    if policy_number:
        leads = leads.filter(policy_number__icontains=policy_number)
    if start_date:
        leads = leads.filter(risk_start_date__gte=start_date)
    if end_date:
        leads = leads.filter(risk_start_date__lte=end_date)
    if insurance_company:
        leads = leads.filter(insurance_company=insurance_company)
    if policy_type:
        leads = leads.filter(policy_type=policy_type)
    if vehicle_type:
        leads = leads.filter(vehicle_type=vehicle_type)
    if lead_type:
        leads = leads.filter(lead_type=lead_type)

    if lead_type == 'MOTOR' and motor_type:
        leads = leads.filter(vehicle_type=motor_type)

     # Example logic for upcoming renewals (next 30 days)
    upcoming_renewals = request.GET.get('upcoming_renewals')

    if upcoming_renewals:
        try:
            today = datetime.today().date()
            days = int(upcoming_renewals)
            target_date = today + timedelta(days=days)

        # Range from today to target_date
            leads = leads.filter(risk_start_date__range=[today, target_date])
        except ValueError:
            pass  # Invalid number of days (safe fallback)
    
   
    # Get unique dropdown values
    sales_managers = Users.objects.filter(role_id=3).values('first_name','first_name', 'last_name').distinct()
      
    agents = Users.objects.filter(role_id=4).values_list('user_name', flat=True)
    insurance_companies = Leads.objects.values_list('insurance_company', flat=True).distinct().exclude(insurance_company__isnull=True).exclude(insurance_company__exact='')
    policy_types = Leads.objects.values_list('policy_type', flat=True).distinct().exclude(policy_type__isnull=True).exclude(policy_type__exact='')
    vehicle_types = Leads.objects.values_list('vehicle_type', flat=True).distinct().exclude(vehicle_type__isnull=True).exclude(vehicle_type__exact='')
   

    # Sorting
    if shorting == 'name_asc':
        leads = leads.order_by('name_as_per_pan')
    elif shorting == 'name_desc':
        leads = leads.order_by('-name_as_per_pan')
    elif shorting == 'recently_added':
        leads = leads.order_by('-created_at')
    elif shorting == 'recently_updated':
        leads = leads.order_by('-updated_at')
    else:
        leads = leads.order_by('-created_at')  # Default sort

    # Count
    total_leads = leads.count()
    motor_leads = Leads.objects.filter(lead_type='MOTOR').count()
    health_leads = Leads.objects.filter(lead_type='HEALTH').count()
    term_leads = Leads.objects.filter(lead_type='TERM').count()

    # Pagination
    paginator = Paginator(leads, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'leads/index.html', {
        'page_obj': page_obj,
        'total_leads': total_leads,
        'per_page': per_page,
        'search_field': search_field,
        'search_query': search_query,
        'shorting': shorting,  # Pass to template to retain selected option
        'motor_leads': motor_leads,
        'health_leads': health_leads,
        'term_leads': term_leads,
        'sales_managers': sales_managers,
        'selected_sales_manager': sales_manager,
        'agents': agents,
        'selected_agent': agent_name,
        'insurance_companies': insurance_companies,
        'policy_types': policy_types,
        'vehicle_types': vehicle_types,
    })

    
def viewlead(request, lead_id):
    if request.user.is_authenticated:
        leads = Leads.objects.all()
        return render(request, 'leads/index.html', {
            'leads': leads,})  # Pass leads to the template
    else:
        return redirect('login')
    
def healthLead(request):
    if request.user.is_authenticated:

        leads = Leads.objects.all()
        # Count
        total_leads = leads.count()
        motor_leads = Leads.objects.filter(lead_type='MOTOR').count()
        health_leads = Leads.objects.filter(lead_type='HEALTH').count()
        term_leads = Leads.objects.filter(lead_type='TERM').count()

        # Get filter inputs
        lead_id = request.GET.get('lead_id', '')
        name = request.GET.get('name_as_per_pan', '')
        pan = request.GET.get('pan_card_number', '')
        email = request.GET.get('email_address', '')
        mobile = request.GET.get('mobile_number', '')
        sales_manager = request.GET.get('sales_manager', '')
        agent_name = request.GET.get('agent_name', '')
        policy_number = request.GET.get('policy_number', '')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        insurance_company = request.GET.get('insurance_company', '')
        policy_type = request.GET.get('policy_type', '')
        vehicle_type = request.GET.get('vehicle_type', '')
        upcoming_renewals = request.GET.get('upcoming_renewals', '')
        lead_type = request.GET.get('lead_type')
        motor_type = request.GET.get('motor_type')

    #today = datetime.today().date()
    #after_30_days = today + timedelta(days=30)
    # Apply filters
        if lead_id:
            leads = leads.filter(lead_id__icontains=lead_id)
        if name:
            leads = leads.filter(name_as_per_pan__icontains=name)
        if pan:
            leads = leads.filter(pan_card_number__icontains=pan)
        if email:
            leads = leads.filter(email_address__icontains=email)
        if mobile:
            leads = leads.filter(mobile_number__icontains=mobile)
        if sales_manager:
            leads = leads.filter(sales_manager__user_name=sales_manager)
        if agent_name:
            leads = leads.filter(agent_name=agent_name)
        if policy_number:
            leads = leads.filter(policy_number__icontains=policy_number)
        if start_date:
            leads = leads.filter(risk_start_date__gte=start_date)
        if end_date:
            leads = leads.filter(risk_start_date__lte=end_date)
        if insurance_company:
            leads = leads.filter(insurance_company=insurance_company)
        if policy_type:
            leads = leads.filter(policy_type=policy_type)
        if vehicle_type:
            leads = leads.filter(vehicle_type=vehicle_type)
        if lead_type:
            leads = leads.filter(lead_type=lead_type)

        if lead_type == 'MOTOR' and motor_type:
            leads = leads.filter(vehicle_type=motor_type)

     # Example logic for upcoming renewals (next 30 days)
        upcoming_renewals = request.GET.get('upcoming_renewals')

        if upcoming_renewals:
            try:
                today = datetime.today().date()
                days = int(upcoming_renewals)
                target_date = today + timedelta(days=days)

        # Range from today to target_date
                leads = leads.filter(risk_start_date__range=[today, target_date])
            except ValueError:
                pass  # Invalid number of days (safe fallback)
    
   
        # Get unique dropdown values
        sales_managers = Users.objects.filter(role_id=3).values('first_name','first_name', 'last_name').distinct()
      
        agents = Users.objects.filter(role_id=4).values_list('user_name', flat=True)
        insurance_companies = Leads.objects.values_list('insurance_company', flat=True).distinct().exclude(insurance_company__isnull=True).exclude(insurance_company__exact='')
        policy_types = Leads.objects.values_list('policy_type', flat=True).distinct().exclude(policy_type__isnull=True).exclude(policy_type__exact='')
        vehicle_types = Leads.objects.values_list('vehicle_type', flat=True).distinct().exclude(vehicle_type__isnull=True).exclude(vehicle_type__exact='')
   

        return render(request, 'leads/health-lead.html',{
            'leads': leads,
            'total_leads':total_leads,
            'motor_leads': motor_leads,
            'health_leads': health_leads,
            'term_leads': term_leads,
            'sales_managers': sales_managers,
            'selected_sales_manager': sales_manager,
            'agents': agents,
            'selected_agent': agent_name,
            'insurance_companies': insurance_companies,
            'policy_types': policy_types,
            'vehicle_types': vehicle_types,
            
        })  
    else:
        return redirect('login')
    
def termlead(request):
    if request.user.is_authenticated:

        leads = Leads.objects.all()
        # Count
        total_leads = leads.count()
        motor_leads = Leads.objects.filter(lead_type='MOTOR').count()
        health_leads = Leads.objects.filter(lead_type='HEALTH').count()
        term_leads = Leads.objects.filter(lead_type='TERM').count()

        # Get filter inputs
        lead_id = request.GET.get('lead_id', '')
        name = request.GET.get('name_as_per_pan', '')
        pan = request.GET.get('pan_card_number', '')
        email = request.GET.get('email_address', '')
        mobile = request.GET.get('mobile_number', '')
        sales_manager = request.GET.get('sales_manager', '')
        agent_name = request.GET.get('agent_name', '')
        policy_number = request.GET.get('policy_number', '')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        insurance_company = request.GET.get('insurance_company', '')
        policy_type = request.GET.get('policy_type', '')
        vehicle_type = request.GET.get('vehicle_type', '')
        upcoming_renewals = request.GET.get('upcoming_renewals', '')
        lead_type = request.GET.get('lead_type')
        motor_type = request.GET.get('motor_type')

    #today = datetime.today().date()
    #after_30_days = today + timedelta(days=30)
    # Apply filters
        if lead_id:
            leads = leads.filter(lead_id__icontains=lead_id)
        if name:
            leads = leads.filter(name_as_per_pan__icontains=name)
        if pan:
            leads = leads.filter(pan_card_number__icontains=pan)
        if email:
            leads = leads.filter(email_address__icontains=email)
        if mobile:
            leads = leads.filter(mobile_number__icontains=mobile)
        if sales_manager:
            leads = leads.filter(sales_manager__user_name=sales_manager)
        if agent_name:
            leads = leads.filter(agent_name=agent_name)
        if policy_number:
            leads = leads.filter(policy_number__icontains=policy_number)
        if start_date:
            leads = leads.filter(risk_start_date__gte=start_date)
        if end_date:
            leads = leads.filter(risk_start_date__lte=end_date)
        if insurance_company:
            leads = leads.filter(insurance_company=insurance_company)
        if policy_type:
            leads = leads.filter(policy_type=policy_type)
        if vehicle_type:
            leads = leads.filter(vehicle_type=vehicle_type)
        if lead_type:
            leads = leads.filter(lead_type=lead_type)

        if lead_type == 'MOTOR' and motor_type:
            leads = leads.filter(vehicle_type=motor_type)

     # Example logic for upcoming renewals (next 30 days)
        upcoming_renewals = request.GET.get('upcoming_renewals')

        if upcoming_renewals:
            try:
                today = datetime.today().date()
                days = int(upcoming_renewals)
                target_date = today + timedelta(days=days)

        # Range from today to target_date
                leads = leads.filter(risk_start_date__range=[today, target_date])
            except ValueError:
                pass  # Invalid number of days (safe fallback)
    
   
        # Get unique dropdown values
        sales_managers = Users.objects.filter(role_id=3).values('first_name','first_name', 'last_name').distinct()
        agents = Users.objects.filter(role_id=4).values_list('user_name', flat=True)
        insurance_companies = Leads.objects.values_list('insurance_company', flat=True).distinct().exclude(insurance_company__isnull=True).exclude(insurance_company__exact='')
        policy_types = Leads.objects.values_list('policy_type', flat=True).distinct().exclude(policy_type__isnull=True).exclude(policy_type__exact='')
        vehicle_types = Leads.objects.values_list('vehicle_type', flat=True).distinct().exclude(vehicle_type__isnull=True).exclude(vehicle_type__exact='')
   


        return render(request, 'leads/term-lead.html',
        {
            'leads': leads,
            'total_leads':total_leads,
            'motor_leads': motor_leads,
            'health_leads': health_leads,
            'term_leads': term_leads,
            'sales_managers': sales_managers,
            'selected_sales_manager': sales_manager,
            'agents': agents,
            'selected_agent': agent_name,
            'insurance_companies': insurance_companies,
            'policy_types': policy_types,
            'vehicle_types': vehicle_types,
        })  # Pass leads to the template
    else:
        return redirect('login')

from datetime import datetime


def create_or_edit_lead(request, lead_id=None):
    if not request.user.is_authenticated:
        return redirect('login')

    customers = QuotationCustomer.objects.all()
    source_leads = SourceMaster.objects.filter(status=True).order_by('source_name')
    referrals = Referral.objects.all()
    lead = None

    if request.method == "GET":
        return render(request, 'leads/create.html', {
            'lead': lead,
            'referrals': referrals,
            'customers': customers
        })

    elif request.method == "POST":
        # Step 1: Collect form inputs
        mobile_number = request.POST.get("mobile_number", "").strip()
        pan_card_number = request.POST.get("pan_card_number", "").strip()
        registration_number = request.POST.get("registration_number", "").strip()
        name_as_per_pan = request.POST.get("name_as_per_pan", "").strip()
        email_address = request.POST.get("email_address", "").strip()

        # ... other form fields you want to collect manually ...

        # Step 2: Try to find matching policy
        matching_policy = PolicyInfo.objects.filter(
            Q(insured_mobile=mobile_number) |
            Q(insured_pan=pan_card_number) |
            Q(policy__registration_number=registration_number)
        ).order_by('-created_at').first()

        # Step 3: Prepare the lead instance
        lead = Leads(
            #lead_id=generate_unique_lead_id(),  # You can implement your logic for this
            mobile_number=mobile_number,
            pan_card_number=pan_card_number,
            registration_number=registration_number,
            email_address=email_address,
            name_as_per_pan=name_as_per_pan,
            # other manual fields...
        )

        # Step 4: If matching policy found, copy fields
        if matching_policy:
            lead.insurance_company = matching_policy.insurance_company
            lead.policy_number = matching_policy.policy_number
            lead.policy_date = matching_policy.policy_issue_date
            lead.policy_type = matching_policy.policy_type
            lead.make_and_model = matching_policy.policy.make_and_model if hasattr(matching_policy.policy, 'make_and_model') else None
            lead.fuel_type = matching_policy.fuel_type
            lead.manufacturing_year = matching_policy.policy.manufacturing_year if hasattr(matching_policy.policy, 'manufacturing_year') else None
            lead.sum_insured = matching_policy.sum_insured
            lead.od_premium = matching_policy.od_premium
            lead.tp_premium = matching_policy.tp_premium
            lead.net_premium = matching_policy.net_premium
            lead.gross_premium = matching_policy.gross_premium
            lead.agent_name = matching_policy.pos_name
            lead.sales_manager = matching_policy.supervisor_name
            # Add more as needed

        # Step 5: Save the lead
        lead.save()

        # Step 6: Redirect or render success
        messages.success(request, "Lead created successfully!")
        return redirect("leads_list")  # Update this to your actual lead list URL

"""def create_or_edit_lead(request, lead_id=None):
    if not request.user.is_authenticated:
        return redirect('login')
    
    customers = QuotationCustomer.objects.all()
    source_leads = SourceMaster.objects.filter(status=True).order_by('source_name')
    print("Loaded sources:", list(source_leads))

    lead = None
    referrals = Referral.objects.all()

    #if lead_id:
        #lead = get_object_or_404(Leads, id=lead_id)
    
    
    if request.method == "GET":
        return render(request, 'leads/create.html', {
            'lead': lead,
            'referrals': referrals,
            'customers': customers
        })


    elif request.method == "POST":
        mobile_number = request.POST.get("mobile_number", "").strip()
        email_address = request.POST.get("email_address", "").strip()
        quote_date = request.POST.get("quote_date", None)
        name_as_per_pan = request.POST.get("name_as_per_pan", "").strip()
        pan_card_number = request.POST.get("pan_card_number", "").strip() or None


        source_leads_id = request.POST.get("source_leads", "").strip() or None
        source_leads = SourceMaster.objects.get(id=source_leads_id) if source_leads_id else None
        
        # ✅ Handle date_of_birth safely
        date_of_birth_str = request.POST.get("date_of_birth", "").strip()
        date_of_birth = None
        if date_of_birth_str:
            try:
                date_of_birth = datetime.strptime(date_of_birth_str, "%Y-%m-%d").date()
            except ValueError:
                messages.error(request, "Invalid date format for Date of Birth. Please use YYYY-MM-DD.")
                return redirect(request.path)

        state = request.POST.get("state", "").strip()
        city = request.POST.get("city", "").strip()
        pincode = request.POST.get("pincode", "").strip()
        address = request.POST.get("address", "").strip()
        lead_source = request.POST.get("lead_source", "").strip()
        referral_by = request.POST.get("referral_by", "").strip()
        if lead_source != 'referral_partner':
            referral_by = ''
        lead_description = request.POST.get("lead_description", "").strip()
        # lead_type = request.POST.get("lead_type", "MOTOR").strip()
        lead_type = request.POST.get("lead_type", "MOTOR").strip()

        if lead_type == "MOTOR":
            registration_number = request.POST.get("registration_number", "").strip()
            vehicle_type = request.POST.get("vehicle_type", "").strip()

        else:
           registration_number = ""
        
        status = request.POST.get("status", "new").strip()
        
        if lead:
            lead.mobile_number = mobile_number
            lead.email_address = email_address
            lead.quote_date = quote_date
            lead.name_as_per_pan = name_as_per_pan
            lead.pan_card_number = pan_card_number
            lead.date_of_birth = date_of_birth
            lead.state = state
            lead.city = city
            lead.pincode = pincode
            lead.address = address
            lead.lead_description = lead_description
            lead.lead_type = lead_type
            lead.registration_number = registration_number  ## vehicle no.
            lead.vehicle_type = vehicle_type
            lead.lead_source = lead_source
            lead.referral_by = referral_by
            lead.status = status
            lead.updated_at = now()
            lead.save()
            messages.success(request, f"Lead updated successfully! Lead ID: {lead.lead_id}")
        else:
            Leads.objects.create(
                mobile_number=mobile_number,
                email_address=email_address,
                quote_date=quote_date,
                name_as_per_pan=name_as_per_pan,
                pan_card_number=pan_card_number,
                date_of_birth=date_of_birth,
                state=state,
                city=city,
                pincode=pincode,
                address=address,
                lead_description=lead_description,
                lead_type=lead_type,
                registration_number=registration_number,
                vehicle_type=vehicle_type,
                lead_source=lead_source,
                referral_by=referral_by,
                status=status,
                created_by=request.user.id,
                created_at=now(),
                updated_at=now()
            )
            messages.success(request, f"Lead created successfully!")

        return redirect("leads-mgt")"""

"""def bulk_upload_leads(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            messages.error(request, "Please select a file to upload.")
            return redirect('leads-mgt')

        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, "Only .xlsx files are supported.")
            return redirect('leads-mgt')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            # Get latest lead_id once before loop
            latest_lead = Leads.objects.aggregate(Max('lead_id'))['lead_id__max']
            start_num = 1
            if latest_lead:
                match = re.search(r'L(\d+)', latest_lead)
                if match:
                    start_num = int(match.group(1)) + 1

            inserted = 0
            duplicate_data_found = False

            for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) < 13 or any(cell is None for cell in row[:13]):
                    logger.warning(f"Row {index} skipped: Incomplete data.")
                    continue

                # Extract values
                mobile_number = str(row[0]).strip()
                email_address = str(row[1]).strip()
                quote_date = row[2]
                name_as_per_pan = str(row[3]).strip()
                pan_card_number = str(row[4]).strip().upper()
                date_of_birth = row[5]
                state = str(row[6]).strip()
                city = str(row[7]).strip()
                pincode = str(row[8]).strip()
                lead_source = str(row[9]).strip()
                address = str(row[10]).strip()
                lead_description = str(row[11]).strip()
                lead_type = str(row[12]).strip()

                # VALIDATION
                if not re.fullmatch(r'[6-9]\d{9}', mobile_number):
                    logger.warning(f"Row {index} skipped: Invalid mobile number - {mobile_number}")
                    continue

                if not re.match(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$", email_address):
                    logger.warning(f"Row {index} skipped: Invalid email - {email_address}")
                    continue

                if not re.match(r'^[A-Z]{5}[0-9]{4}[A-Z]$', pan_card_number):
                    logger.warning(f"Row {index} skipped: Invalid PAN - {pan_card_number}")
                    continue

                try:
                    dob = parser.parse(str(date_of_birth)).date()
                except Exception as e:
                    logger.warning(f"Row {index} skipped: Invalid DOB - {date_of_birth}")
                    continue

                # Duplicate Check
                if Leads.objects.filter(
                    Q(mobile_number=mobile_number) |
                    Q(email_address=email_address) |
                    Q(pan_card_number=pan_card_number)
                ).exists():
                    logger.info(f"Row {index} skipped: Duplicate entry - Mobile: {mobile_number}, Email: {email_address}, PAN: {pan_card_number}")
                    duplicate_data_found = True
                    continue

                # Generate lead_id
                lead_id = f"L{start_num:05d}"
                start_num += 1

                # Insert into DB
                try:
                    Leads.objects.create(
                        lead_id=lead_id,
                        mobile_number=mobile_number,
                        email_address=email_address,
                        quote_date=quote_date,
                        name_as_per_pan=name_as_per_pan,
                        pan_card_number=pan_card_number,
                        date_of_birth=dob,
                        state=state,
                        city=city,
                        pincode=pincode,
                        lead_source=lead_source,
                        address=address,
                        lead_description=lead_description,
                        lead_type=lead_type,
                    )
                    inserted += 1

                except Exception as e:
                    logger.error(f"Row {index} error: {e}")
                    continue

            # Final message
            if inserted > 0:
                messages.success(request, f"{inserted} leads uploaded successfully.")
            elif duplicate_data_found:
                messages.warning(request, "No new leads were inserted. All records were duplicates.")
            else:
                messages.info(request, "No valid data found in Excel file!")

            return redirect('leads-mgt')

        except Exception as e:
            logger.error(f"Error reading Excel: {e}")
            messages.error(request, f"Error processing file: {e}")
            return redirect('leads-mgt')

    return render(request, 'leads/bulk_upload.html')"""


"""def bulk_upload_leads(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            messages.error(request, "Please select a file to upload.")
            return redirect('leads-mgt')

        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, "Only .xlsx files are supported.")
            return redirect('leads-mgt')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            latest_lead = Leads.objects.aggregate(Max('lead_id'))['lead_id__max']
            start_num = 1
            if latest_lead:
                match = re.search(r'L(\d+)', latest_lead)
                if match:
                    start_num = int(match.group(1)) + 1

            inserted = 0
            duplicate_data_found = False

            for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) < 29 or any(cell is None for cell in row[:13]):
                    logger.warning(f"Row {index} skipped: Incomplete data.")
                    continue

                try:
                    mobile_number = str(row[0]).strip()
                    email_address = str(row[1]).strip()
                    quote_date = row[2]
                    name_as_per_pan = str(row[3]).strip()
                    pan_card_number = str(row[4]).strip().upper()
                    date_of_birth = row[5]
                    state = str(row[6]).strip()
                    city = str(row[7]).strip()
                    pincode = str(row[8]).strip()
                    lead_source = str(row[9]).strip()
                    address = str(row[10]).strip()
                    lead_description = str(row[11]).strip()
                    lead_type = str(row[12]).strip()

                    # New fields
                    policy_date = row[13]
                    sales_manager = str(row[14]).strip()
                    agent_name = str(row[15]).strip()
                    insurance_company = str(row[16]).strip()
                    policy_type = str(row[17]).strip()
                    policy_number = str(row[18]).strip()
                    vehicle_type = str(row[19]).strip()
                    make_and_model = str(row[20]).strip()
                    fuel_type = str(row[21]).strip()
                    registration_number = str(row[22]).strip()
                    manufacturing_year = row[23]
                    sum_insured = row[24]
                    ncb = row[25]
                    od_premium = row[26]
                    tp_premium = row[27]
                    risk_start_date = row[28]

                    # VALIDATIONS
                    if not re.fullmatch(r'[6-9]\d{9}', mobile_number):
                        logger.warning(f"Row {index} skipped: Invalid mobile number - {mobile_number}")
                        continue

                    if not re.match(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$", email_address):
                        logger.warning(f"Row {index} skipped: Invalid email - {email_address}")
                        continue

                    if not re.match(r'^[A-Z]{5}[0-9]{4}[A-Z]$', pan_card_number):
                        logger.warning(f"Row {index} skipped: Invalid PAN - {pan_card_number}")
                        continue

                    try:
                        dob = parser.parse(str(date_of_birth)).date()
                    except Exception as e:
                        logger.warning(f"Row {index} skipped: Invalid DOB - {date_of_birth}")
                        continue

                    try:
                        policy_date = parser.parse(str(policy_date)).date() if policy_date else None
                    except:
                        policy_date = None

                    try:
                        risk_start_date = parser.parse(str(risk_start_date)).date() if risk_start_date else None
                    except:
                        risk_start_date = None

                    # Duplicate check
                    if Leads.objects.filter(
                        Q(mobile_number=mobile_number) |
                        Q(email_address=email_address) |
                        Q(pan_card_number=pan_card_number)
                    ).exists():
                        logger.info(f"Row {index} skipped: Duplicate entry.")
                        duplicate_data_found = True
                        continue

                    # Create new lead_id
                    lead_id = f"L{start_num:05d}"
                    start_num += 1

                    # Insert into DB
                    Leads.objects.create(
                        lead_id=lead_id,
                        mobile_number=mobile_number,
                        email_address=email_address,
                        quote_date=quote_date,
                        name_as_per_pan=name_as_per_pan,
                        pan_card_number=pan_card_number,
                        date_of_birth=dob,
                        state=state,
                        city=city,
                        pincode=pincode,
                        lead_source=lead_source,
                        address=address,
                        lead_description=lead_description,
                        lead_type=lead_type,
                        policy_date=policy_date,
                        sales_manager=sales_manager,
                        agent_name=agent_name,
                        insurance_company=insurance_company,
                        policy_type=policy_type,
                        policy_number=policy_number,
                        vehicle_type=vehicle_type,
                        make_and_model=make_and_model,
                        fuel_type=fuel_type,
                        registration_number=registration_number,
                        manufacturing_year=manufacturing_year,
                        sum_insured=sum_insured,
                        ncb=ncb,
                        od_premium=od_premium,
                        tp_premium=tp_premium,
                        net_premium=(od_premium or 0) + (tp_premium or 0),
                        gross_premium=(od_premium or 0) + (tp_premium or 0),
                        risk_start_date=risk_start_date
                    )
                    inserted += 1

                except Exception as e:
                    logger.error(f"Row {index} error: {e}")
                    continue

            # Final messages
            if inserted > 0:
                messages.success(request, f"{inserted} leads uploaded successfully.")
            elif duplicate_data_found:
                messages.warning(request, "No new leads were inserted. All records were duplicates.")
            else:
                messages.info(request, "No valid data found in Excel file!")

            return redirect('leads-mgt')

        except Exception as e:
            logger.error(f"Error reading Excel: {e}")
            messages.error(request, f"Error processing file: {e}")
            return redirect('leads-mgt')

    return render(request, 'leads/bulk_upload.html')"""


@login_required
def bulk_upload_leads(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        camp_name = request.POST.get("camp_name")

        if not excel_file.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, "Only Excel files (.xlsx, .xls) are allowed!")
            return redirect('bulk-upload-leads')

        if not camp_name:
            messages.error(request, "Campaign Name is mandatory.")
            return redirect('bulk-upload-leads')


        # Save the file record
        instance = LeadUploadExcel.objects.create(
            file=excel_file,
            file_name=excel_file.name,
            file_url=excel_file.name,
            campaign_name=camp_name,
            created_by=request.user  # assuming request.user is correct
        )

        # Trigger background task
        async_task("empPortal.taskz.process_lead_excel.process_lead_excel", instance.id)

        messages.success(request, "File uploaded. It will be processed in the background.")
        return redirect("leads-mgt")

    return render(request, "leads/bulk_upload.html")
    