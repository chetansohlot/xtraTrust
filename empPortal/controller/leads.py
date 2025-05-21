# Standard Library Imports
import os
import re
import time
import json
import logging
import zipfile
from io import BytesIO
from datetime import datetime, timedelta
from pprint import pprint
from urllib import request

# Third-Party Imports
import fitz  # PyMuPDF
import openai
import pandas as pd
import openpyxl
from dateutil import parser
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from fastapi import FastAPI, File, UploadFile

# Django Core Imports
from django.conf import settings
from django.db import connection
from django.db.models import Q, F, Value, Max, CharField, OuterRef, Subquery
from django.db.models.functions import Concat, Coalesce
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template import loader
from django.template.loader import render_to_string
from django.core.mail import send_mail, EmailMessage
from django.core.files.storage import FileSystemStorage, default_storage
from django.core.paginator import Paginator
from django.contrib import messages
from django.contrib.auth import (
    authenticate, login, logout,
    update_session_auth_hash,
    hashers as auth_hashers
)
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.utils.timezone import now
from django_q.tasks import async_task

# Local App Imports
from ..models import (
    Users, Roles, Commission, LeadUploadExcel, SourceMaster,
    DocumentUpload, Branch, Leads, QuotationCustomer,
    PolicyInfo, PolicyDocument
)
from ..forms import DocumentUploadForm
from ..model import (
    State, City, InsuranceType,
    InsuranceCategory, InsuranceProduct, Insurance
)

from empPortal.model import (
    BankDetails, Referral, Partner
)
from empPortal.model.customer import Customer
from empPortal.model.leadActivity import LeadActivity
from empPortal.model.Dispositions import Disposition, SubDisposition
from empPortal.model.LeadDisposition import LeadDisposition, LeadDispositionLogs
from empPortal.model.leads import LeadPreviousPolicy
from empPortal.model.policyTypes import PolicyType
from empPortal.model.vehicleTypes import VehicleType

logger = logging.getLogger(__name__)

def dictfetchall(cursor):
    "Returns all rows from a cursor as a dict"
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]

def index(request):
    if not request.user.is_authenticated:
        return redirect('login')

    per_page = request.GET.get('per_page', 10)
    search_field = request.GET.get('search_field', '')
    search_query = request.GET.get('search_query', '')
    global_search = request.GET.get('global_search', '').strip()
    shorting = request.GET.get('shorting', '')  # Get sorting preference

    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10

    role_id = request.user.role_id
    user_id = request.user.id

    leads = Leads.objects.filter(status=True)
    
    if role_id == 2:  # Management
        leads = leads

    elif role_id == 3:  # Branch Manager
        managers = Users.objects.filter(role_id=5, senior_id=user_id)
        team_leaders = Users.objects.filter(role_id=6, senior_id__in=managers.values_list('id', flat=True))
        relationship_managers = Users.objects.filter(role_id=7, senior_id__in=team_leaders.values_list('id', flat=True))

        user_ids = list(managers.values_list('id', flat=True)) + \
                list(team_leaders.values_list('id', flat=True)) + \
                list(relationship_managers.values_list('id', flat=True)) + \
                [user_id]

        leads = leads.filter(Q(created_by_id__in=user_ids) | Q(assigned_to_id__in=user_ids))
    elif role_id == 4:  # Agent
        leads = leads.filter(Q(created_by_id=user_id) | Q(assigned_to_id=user_id))

    elif role_id == 5:  # Manager
        team_leaders = Users.objects.filter(role_id=6, senior_id=user_id)
        relationship_managers = Users.objects.filter(role_id=7, senior_id__in=team_leaders.values_list('id', flat=True))

        team_leader_ids = team_leaders.values_list('id', flat=True)
        rm_ids = relationship_managers.values_list('id', flat=True)
        user_ids = list(team_leader_ids) + list(rm_ids) + [user_id]

        leads = leads.filter(
            Q(created_by_id__in=user_ids) | Q(assigned_to_id__in=user_ids)
        )

    elif role_id == 6:  # Team Leader
        relationship_managers = Users.objects.filter(role_id=7, senior_id=user_id)
        rm_ids = relationship_managers.values_list('id', flat=True)
        user_ids = list(rm_ids) + [user_id]
        leads = leads.filter(
            Q(created_by_id__in=user_ids) | Q(assigned_to_id__in=user_ids)
        )

    elif role_id == 7:  # Relationship Manager
        leads =leads.filter(
            Q(created_by_id=user_id) | Q(assigned_to_id=user_id)
        )

    else:
        leads =leads

    # Global search
    if global_search:
        leads = leads.filter(
            Q(name_as_per_pan__icontains=global_search) |
            Q(email_address__icontains=global_search) |
            Q(mobile_number__icontains=global_search) |
            Q(pan_card_number__icontains=global_search) |
            Q(state__icontains=global_search) |
            Q(city__icontains=global_search) |
            Q(lead_id__icontains=global_search)
        )

    # Field-specific search
    if search_field and search_query:
        filter_args = {f"{search_field}__icontains": search_query}
        leads = leads.filter(**filter_args)
    
    # Get filter inputs
    lead_id = request.GET.get('lead_id', '')
    name = request.GET.get('name_as_per_pan', '')
    identity_no = request.GET.get('pan_card_number', '')
    email = request.GET.get('email_address', '')
    mobile = request.GET.get('mobile_number', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    branch = request.GET.get('branch','')
    sales_manager = request.GET.get('sales_manager', '')
    sales_teamleader = request.GET.get('sales_teamleader', '')
    sales_rm = request.GET.get('sales_rm', '')
    agent_name = request.GET.get('agent_name', '')
    policy_number = request.GET.get('policy_number', '')
    
    insurance_company = request.GET.get('insurance_company', '')
    policy_type = request.GET.get('policy_type', '')
    vehicle_type = request.GET.get('vehicle_type', '')
    upcoming_renewals = request.GET.get('upcoming_renewals', '')
    lead_type = request.GET.get('lead_type')
    motor_type = request.GET.get('motor_type')

    # Collect filter inputs
    filters_applied = any([
        lead_id, name, identity_no, email, mobile, start_date, end_date, 
        branch, sales_manager, sales_teamleader,sales_rm,
        agent_name, policy_number,
        insurance_company,
        policy_type, vehicle_type, upcoming_renewals,
        lead_type, motor_type
    ])

    allowed_roles = Roles.objects.filter(roleDepartment=1)

    # Apply filters
    if lead_id:
        leads = leads.filter(lead_id__icontains=lead_id)
    if name:
        leads = leads.filter(name_as_per_pan__icontains=name)
    if identity_no:
        leads = leads.filter(lead_customer_identity_no__icontains=identity_no)
    if email:
        leads = leads.filter(email_address__icontains=email)
    if mobile:
        leads = leads.filter(mobile_number__icontains=mobile)
    if start_date:
        leads = leads.filter(created_at__gte=start_date)
    if end_date:
        leads = leads.filter(created_at__lte=end_date)
    if branch:
        leads = leads.filter(branch=branch)
    if sales_manager:
        leads = leads.filter(assigned_manager_id=sales_manager)
    if sales_teamleader:
        leads = leads.filter(assigned_teamleader_id=sales_teamleader)
    if sales_rm:
        leads = leads.filter(assigned_to_id=sales_rm)
    if agent_name:
        leads = leads.filter(agent_name=agent_name)
    if policy_number:
        leads = leads.filter(registration_number__icontains=policy_number)
    if insurance_company:
        lp_subquery = LeadPreviousPolicy.objects.filter(
            lead=OuterRef('id')
        ).values('insurance_company_id')[:1]

        leads = leads.annotate(
            policy_insurance_company=Subquery(lp_subquery)
        ).filter(
            Q(lead_insurance_product_id=32, policy_insurance_company=insurance_company) |
            Q(~Q(lead_insurance_product_id=32), insurance_company=insurance_company)
        )
    if policy_type:
        lp_subquery = LeadPreviousPolicy.objects.filter(
            lead=OuterRef('id')
        ).values('policy_type_id')[:1]

        leads = leads.annotate(
            policy_policy_type=Subquery(lp_subquery)
        ).filter(
            Q(lead_insurance_product_id=32, policy_policy_type=policy_type) |
            Q(~Q(lead_insurance_product_id=32), policy_type=policy_type)
        )
    if vehicle_type:
        lp_subquery = LeadPreviousPolicy.objects.filter(
            lead=OuterRef('id')
        ).values('vehicle_type_id')[:1]

        leads = leads.annotate(
            policy_vehicle_type=Subquery(lp_subquery)
        ).filter(
            Q(lead_insurance_product_id=32, policy_vehicle_type=vehicle_type) |
            Q(~Q(lead_insurance_product_id=32), vehicle_type=vehicle_type)
        )
    
    upcoming_renewals = request.GET.get('upcoming_renewals')

    if upcoming_renewals:
        try:
            today = datetime.today().date()
            days = int(upcoming_renewals)
            target_date = today + timedelta(days=days)

            leads = leads.filter(risk_start_date__range=[today, target_date])
        except ValueError:
            pass  # Invalid number of days (safe fallback)
    
   
    sales_managers = Users.objects.filter(
        role_id=3,
        role__in=allowed_roles
        ).values('first_name', 'last_name').distinct() 
    agents = Users.objects.filter(role_id=4).values_list('user_name', flat=True)
    insurance_companies = leads.values_list('insurance_company', flat=True).distinct().exclude(insurance_company__isnull=True).exclude(insurance_company__exact='')
    policy_types = leads.values_list('policy_type', flat=True).distinct().exclude(policy_type__isnull=True).exclude(policy_type__exact='')
    vehicle_types = leads.values_list('vehicle_type', flat=True).distinct().exclude(vehicle_type__isnull=True).exclude(vehicle_type__exact='')
   
    # Sorting
    if shorting == 'name_asc':
        leads = leads.order_by('name_as_per_pan')
    elif shorting == 'name_desc':
        leads = leads.order_by('-name_as_per_pan')
    elif shorting == 'recently_added':
        leads = leads.order_by('-created_at')
    elif shorting == 'recently_updated':
        leads = leads.order_by('-updated_at')
    else:
        leads = leads.order_by('-created_at')  # Default sort

    if request.GET.get('export') == '1':
        if filters_applied:
            return export_leads_to_excel(leads)
        else:
            messages.warning(request, "Please select at least one filter to export data.")
            return redirect('leads-mgt')
    
   
    total_leads = leads.count()  
    motor_leads = leads.filter(lead_type='MOTOR').count()
    health_leads = leads.filter(lead_type='HEALTH').count()
    term_leads = leads.filter(lead_type='TERM').count()

    # Pagination
    paginator = Paginator(leads, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    branch = Branch.objects.filter(status="Active")
    policy_type_list = PolicyType.objects.filter(status=1)
    vehicle_type_list = VehicleType.objects.filter(status=1)
    insurance_companies = Insurance.objects.filter(active="active")
    return render(request, 'leads/index.html', {
        'page_obj': page_obj,
        'total_leads': total_leads,
        'per_page': per_page,
        'search_field': search_field,
        'search_query': search_query,
        'shorting': shorting,  # Pass to template to retain selected option
        'motor_leads': motor_leads,
        'health_leads': health_leads,
        'term_leads': term_leads,
        'sales_managers': sales_managers,
        'selected_sales_manager': sales_manager,
        'agents': agents,
        'selected_agent': agent_name,
        'insurance_companies': insurance_companies,
        'vehicle_types': vehicle_types,
        'leads': leads,
        'branchs': branch,
        'policy_type_list': policy_type_list,
        'vehicle_type_list': vehicle_type_list
    })

def export_leads_to_excel(leads_queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Step 1: Check if any lead has a policy
    has_policy_data = False
    for lead in leads_queryset:
        policy = PolicyDocument.objects.filter(vehicle_number=lead.registration_number).order_by('-created_at').first()
        if policy and policy.policy_number:
            has_policy_data = True
            break

    # Step 2: Define headers dynamically
    headers = [
        'S.No', 'Lead ID', 'Name as per PAN', 'Email Address', 'Mobile Number',
        'PAN Card Number', 'Vehicle Number'
    ]
    if has_policy_data:
        headers += [
            'Previous Policy Number', 'Policy Issue Date', 'Policy Expiry Date',
            'Sum Insured', 'Net Insurance', 'Gross Insurance'
        ]
    headers += ['Created Date']

    ws.append(headers)
    ws.freeze_panes = 'A2'

    # Step 3: Write rows
    for index, lead in enumerate(leads_queryset, start=1):
        previous_policy = PolicyDocument.objects.filter(vehicle_number=lead.registration_number).order_by('-created_at').first()

        row = [
            index,
            lead.id,
            lead.name_as_per_pan or '',
            lead.email_address or '',
            lead.mobile_number or '',
            lead.pan_card_number or '',
            lead.registration_number or '',
          
        ]

        if has_policy_data:
            row += [
                previous_policy.policy_number if previous_policy else '',
                previous_policy.policy_issue_date if previous_policy else '',
                previous_policy.policy_expiry_date if previous_policy else '',
                previous_policy.sum_insured if previous_policy else '',
                previous_policy.policy_premium if previous_policy else '',
                previous_policy.policy_total_premium if previous_policy else '',
            ]

        row += [
            lead.created_at.strftime('%Y-%m-%d %H:%M:%S') if lead.created_at else '',
            
        ]
            
        ws.append(row)

    # Step 4: Adjust column widths
    for i, column_cells in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = max_length + 2

    # Step 5: Return Excel response
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    response = HttpResponse(file_stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="filtered_leads.xlsx"'

    return response


def viewlead(request, lead_id):
    if request.user.is_authenticated:
        leads = Leads.objects.all()
        return render(request, 'leads/index.html', {
            'leads': leads,})  # Pass leads to the template
    else:
        return redirect('login')
    
def healthLead(request):
    if request.user.is_authenticated:

        leads = Leads.objects.all()
        # Count
        all_leads = Leads.objects.all()
        total_leads = all_leads.count()  
        motor_leads = Leads.objects.filter(lead_type='MOTOR').count()
        health_leads = Leads.objects.filter(lead_type='HEALTH').count()
        term_leads = Leads.objects.filter(lead_type='TERM').count()

        # Get filter inputs
        lead_id = request.GET.get('lead_id', '')
        name = request.GET.get('name_as_per_pan', '')
        pan = request.GET.get('pan_card_number', '')
        email = request.GET.get('email_address', '')
        mobile = request.GET.get('mobile_number', '')
        sales_manager = request.GET.get('sales_manager', '')
        agent_name = request.GET.get('agent_name', '')
        policy_number = request.GET.get('policy_number', '')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        insurance_company = request.GET.get('insurance_company', '')
        policy_type = request.GET.get('policy_type', '')
        vehicle_type = request.GET.get('vehicle_type', '')
        upcoming_renewals = request.GET.get('upcoming_renewals', '')
        lead_type = request.GET.get('lead_type')
        motor_type = request.GET.get('motor_type')

    #today = datetime.today().date()
    #after_30_days = today + timedelta(days=30)
    # Apply filters
        if lead_id:
            leads = leads.filter(lead_id__icontains=lead_id)
        if name:
            leads = leads.filter(name_as_per_pan__icontains=name)
        if pan:
            leads = leads.filter(pan_card_number__icontains=pan)
        if email:
            leads = leads.filter(email_address__icontains=email)
        if mobile:
            leads = leads.filter(mobile_number__icontains=mobile)
        if sales_manager:
            leads = leads.filter(sales_manager__user_name=sales_manager)
        if agent_name:
            leads = leads.filter(agent_name=agent_name)
        if policy_number:
            leads = leads.filter(registration_number__icontains=policy_number)
        if start_date:
            leads = leads.filter(risk_start_date__gte=start_date)
        if end_date:
            leads = leads.filter(risk_start_date__lte=end_date)
        if insurance_company:
            leads = leads.filter(insurance_company=insurance_company)
        if policy_type:
            leads = leads.filter(policy_type=policy_type)
        if vehicle_type:
            leads = leads.filter(vehicle_type=vehicle_type)
        if lead_type:
            leads = leads.filter(lead_type=lead_type)

        if lead_type == 'MOTOR' and motor_type:
            leads = leads.filter(vehicle_type=motor_type)

     # Example logic for upcoming renewals (next 30 days)
        upcoming_renewals = request.GET.get('upcoming_renewals')

        if upcoming_renewals:
            try:
                today = datetime.today().date()
                days = int(upcoming_renewals)
                target_date = today + timedelta(days=days)

        # Range from today to target_date
                leads = leads.filter(risk_start_date__range=[today, target_date])
            except ValueError:
                pass  # Invalid number of days (safe fallback)
    
   
        # Get unique dropdown values
        sales_managers = Users.objects.filter(role_id=3).values('first_name','first_name', 'last_name').distinct()
      
        agents = Users.objects.filter(role_id=4).values_list('user_name', flat=True)
        insurance_companies = Leads.objects.values_list('insurance_company', flat=True).distinct().exclude(insurance_company__isnull=True).exclude(insurance_company__exact='')
        policy_types = Leads.objects.values_list('policy_type', flat=True).distinct().exclude(policy_type__isnull=True).exclude(policy_type__exact='')
        vehicle_types = Leads.objects.values_list('vehicle_type', flat=True).distinct().exclude(vehicle_type__isnull=True).exclude(vehicle_type__exact='')

         # After filtering leads
        if request.GET.get('export') == '1':
            return export_leads_to_excel(leads)

        return render(request, 'leads/health-lead.html',{
            'leads': leads,
            'total_leads':total_leads,
            'motor_leads': motor_leads,
            'health_leads': health_leads,
            'term_leads': term_leads,
            'sales_managers': sales_managers,
            'selected_sales_manager': sales_manager,
            'agents': agents,
            'selected_agent': agent_name,
            'insurance_companies': insurance_companies,
            'policy_types': policy_types,
            #'vehicle_types': vehicle_types,
            
        })  
    else:
        return redirect('login')
    
def termlead(request):
    if request.user.is_authenticated:

        leads = Leads.objects.all()
        # Count
        all_leads = Leads.objects.all()
        total_leads = all_leads.count()  
        motor_leads = Leads.objects.filter(lead_type='MOTOR').count()
        health_leads = Leads.objects.filter(lead_type='HEALTH').count()
        term_leads = Leads.objects.filter(lead_type='TERM').count()

        # Get filter inputs
        lead_id = request.GET.get('lead_id', '')
        name = request.GET.get('name_as_per_pan', '')
        pan = request.GET.get('pan_card_number', '')
        email = request.GET.get('email_address', '')
        mobile = request.GET.get('mobile_number', '')
        sales_manager = request.GET.get('sales_manager', '')
        agent_name = request.GET.get('agent_name', '')
        policy_number = request.GET.get('policy_number', '')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        insurance_company = request.GET.get('insurance_company', '')
        policy_type = request.GET.get('policy_type', '')
        vehicle_type = request.GET.get('vehicle_type', '')
        upcoming_renewals = request.GET.get('upcoming_renewals', '')
        lead_type = request.GET.get('lead_type')
        motor_type = request.GET.get('motor_type')

    #today = datetime.today().date()
    #after_30_days = today + timedelta(days=30)
    # Apply filters
        if lead_id:
            leads = leads.filter(lead_id__icontains=lead_id)
        if name:
            leads = leads.filter(name_as_per_pan__icontains=name)
        if pan:
            leads = leads.filter(pan_card_number__icontains=pan)
        if email:
            leads = leads.filter(email_address__icontains=email)
        if mobile:
            leads = leads.filter(mobile_number__icontains=mobile)
        if sales_manager:
            leads = leads.filter(sales_manager__user_name=sales_manager)
        if agent_name:
            leads = leads.filter(agent_name=agent_name)
        if policy_number:
            leads = leads.filter(registration_number__icontains=policy_number)
        if start_date:
            leads = leads.filter(risk_start_date__gte=start_date)
        if end_date:
            leads = leads.filter(risk_start_date__lte=end_date)
        if insurance_company:
            leads = leads.filter(insurance_company=insurance_company)
        if policy_type:
            leads = leads.filter(policy_type=policy_type)
        if vehicle_type:
            leads = leads.filter(vehicle_type=vehicle_type)
        if lead_type:
            leads = leads.filter(lead_type=lead_type)

        if lead_type == 'MOTOR' and motor_type:
            leads = leads.filter(vehicle_type=motor_type)

     # Example logic for upcoming renewals (next 30 days)
        upcoming_renewals = request.GET.get('upcoming_renewals')

        if upcoming_renewals:
            try:
                today = datetime.today().date()
                days = int(upcoming_renewals)
                target_date = today + timedelta(days=days)

        # Range from today to target_date
                leads = leads.filter(risk_start_date__range=[today, target_date])
            except ValueError:
                pass  # Invalid number of days (safe fallback)
    
   
        # Get unique dropdown values
        sales_managers = Users.objects.filter(role_id=3).values('first_name','first_name', 'last_name').distinct()
        agents = Users.objects.filter(role_id=4).values_list('user_name', flat=True)
        insurance_companies = Leads.objects.values_list('insurance_company', flat=True).distinct().exclude(insurance_company__isnull=True).exclude(insurance_company__exact='')
        policy_types = Leads.objects.values_list('policy_type', flat=True).distinct().exclude(policy_type__isnull=True).exclude(policy_type__exact='')
        vehicle_types = Leads.objects.values_list('vehicle_type', flat=True).distinct().exclude(vehicle_type__isnull=True).exclude(vehicle_type__exact='')
   


        return render(request, 'leads/term-lead.html',
        {
            'leads': leads,
            'total_leads':total_leads,
            'motor_leads': motor_leads,
            'health_leads': health_leads,
            'term_leads': term_leads,
            'sales_managers': sales_managers,
            'selected_sales_manager': sales_manager,
            'agents': agents,
            'selected_agent': agent_name,
            'insurance_companies': insurance_companies,
            'policy_types': policy_types,
            'vehicle_types': vehicle_types,
        })  # Pass leads to the template
    else:
        return redirect('login')

# For AJAX - Load categories
def load_categories(request):
    type_id = request.GET.get('insurance_type')
    categories = InsuranceCategory.objects.filter(insurance_type_id=type_id).values('id', 'name')
    return JsonResponse(list(categories), safe=False)

# For AJAX - Load products
def load_products(request):
    category_id = request.GET.get('insurance_category')
    products = InsuranceProduct.objects.filter(category_id=category_id).values('id', 'name')
    return JsonResponse(list(products), safe=False)

def create_or_edit_lead(request, lead_id=None):
    if not request.user.is_authenticated:
        return redirect('login')
    
    customers = QuotationCustomer.objects.all()
    source_leads = SourceMaster.objects.filter(status=True).order_by('source_name')
    print("Loaded sources:", list(source_leads))

    lead = None
    referrals = Referral.objects.all()
    partners = Partner.objects.filter(active=True)

    if lead_id:
        lead = get_object_or_404(Leads, id=lead_id)
    else:
        lead = None  
    
    states = State.objects.all()
    
    
    if request.method == "GET":
        return render(request, 'leads/create.html', {
            'lead': lead,
            'referrals': referrals,
            'customers': customers,
            'states': states,
            'partners':partners
        })


    elif request.method == "POST":
        mobile_number = request.POST.get("mobile_number", "").strip()
        email_address = request.POST.get("email_address", "").strip()
        quote_date = request.POST.get("quote_date", None)
        name_as_per_pan = request.POST.get("name_as_per_pan", "").strip()
        pan_card_number = request.POST.get("pan_card_number", "").strip() or None

        source_leads_id = request.POST.get("source_leads", "").strip() or None
        source_leads = SourceMaster.objects.get(id=source_leads_id) if source_leads_id else None
        
        # ✅ Handle date_of_birth safely
        date_of_birth_str = request.POST.get("date_of_birth", "").strip()
        date_of_birth = None
        if date_of_birth_str:
            try:
                date_of_birth = datetime.strptime(date_of_birth_str, "%Y-%m-%d").date()
            except ValueError:
                messages.error(request, "Invalid date format for Date of Birth. Please use YYYY-MM-DD.")
                return redirect(request.path)

        #state = request.POST.get("state", "").strip()
        #city = request.POST.get("city", "").strip()
        state_name = request.POST.get('state')
        city_name = request.POST.get('city')
        pincode = request.POST.get("pincode", "").strip()
        address = request.POST.get("address", "").strip()
        lead_source = request.POST.get("lead_source", "").strip()
        referral_by = request.POST.get("referral_by", "").strip()
        partner_id = request.POST.get("partner_id", "").strip()
        if lead_source != 'referral_partner':
            referral_by = ''
        if lead_source != 'pos_partner':
            partner_id = None
        lead_description = request.POST.get("lead_description", "").strip()
        lead_type = request.POST.get("lead_type", "").strip()

        if lead_type == "MOTOR":
            registration_number = request.POST.get("registration_number", "").strip()
            
            vehicle_type = request.POST.get("vehicle_type", "").strip()

        else:
           registration_number = ""
           vehicle_type = ""  

        # Before creating Lead, try to fetch matching PolicyDocument
        policy_document = None
        if registration_number:
                policy_document = PolicyDocument.objects.filter(vehicle_number=registration_number).first()

                policy_start_date_str = policy_document.policy_start_date if policy_document else None
                try:
                    # Try parsing with both date and time
                    if policy_start_date_str:
                        policy_start_date_obj = datetime.strptime(policy_start_date_str, '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                        # If there's no time component, parse as date only
                        if policy_start_date_str:
                            policy_start_date_obj = datetime.strptime(policy_start_date_str, '%Y-%m-%d')
                    except ValueError:
                        policy_start_date_obj = None

        # Default fields
        insurance_company = None
        policy_number = None
        policy_type = None
        vehicle_type_db = None
        sum_insured = None
        policy_date = None
    
        # If matching policy found, extract details
        if policy_document:
            insurance_company = policy_document.insurance_provider
            policy_number = policy_document.policy_number
            policy_type = policy_document.policy_type
            vehicle_type_db = policy_document.vehicle_type
            sum_insured = policy_document.sum_insured
            policy_date = policy_start_date_obj.strftime('%Y-%m-%d %H:%M:%S')
           
       
        status = request.POST.get("status", "new").strip()
        if lead:
            lead.mobile_number = mobile_number
            lead.email_address = email_address
            lead.quote_date = quote_date
            lead.name_as_per_pan = name_as_per_pan
            lead.pan_card_number = pan_card_number
            lead.date_of_birth = date_of_birth
            lead.state = state_name
            lead.city = city_name
            lead.pincode = pincode
            lead.address = address
            lead.lead_description = lead_description
            lead.lead_type = lead_type
            lead.registration_number = registration_number  ## vehicle no.
            lead.vehicle_type = vehicle_type
            lead.lead_source = lead_source
            lead.referral_by = referral_by
            lead.partner_id = partner_id
            lead.status = status
            lead.updated_at = now()
            lead.insurance_company=insurance_company
            lead.policy_number=policy_number
            lead.policy_type=policy_type
            lead.sum_insured=sum_insured
            lead.risk_start_date=policy_date
            lead.save()
            messages.success(request, f"Lead updated successfully! Lead ID: {lead.lead_id}")
        else:
            new_lead = Leads.objects.create(
                mobile_number=mobile_number,
                email_address=email_address,
                quote_date=quote_date,
                name_as_per_pan=name_as_per_pan,
                pan_card_number=pan_card_number,
                date_of_birth=date_of_birth,
                state=state_name,
                city=city_name,
                pincode=pincode,
                address=address,
                lead_description=lead_description,
                lead_type=lead_type,
                registration_number=registration_number,
                vehicle_type=vehicle_type,
                #vehicle_type = request.POST.get("vehicle_type", None)
                lead_source=lead_source,
                referral_by=referral_by,
                status=status,
                created_by=request.user.id,
                created_at=now(),
                updated_at=now(),
                insurance_company=insurance_company,
                policy_number=policy_number,
                policy_type=policy_type,
                sum_insured=sum_insured,
                risk_start_date = policy_date,
                partner_id = partner_id
                
            )

            # Step 2: Generate lead_id using date + ID
           
            today_str = datetime.today().strftime('%Y%m%d')  # e.g., 20250508
            lead_id = f"{today_str}{new_lead.id}"   

            # Step 3: Save it to the lead
            new_lead.lead_id = lead_id
            new_lead.save()
            
            message = 'New lead is created'
            
            LeadActivity.objects.create(
                lead_id = new_lead.id,
                lead_ref_id = lead_id,
                message = message,
                created_by = request.user.id,
            )
            messages.success(request, f"Lead created successfully!  {lead_id}")

        return redirect("leads-mgt")
    
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from datetime import datetime

@csrf_exempt
def fetch_policy_details(request):
    if request.method == "POST":
        registration_number = request.POST.get("registration_number")
        try:
            policy_document = PolicyDocument.objects.filter(vehicle_number=registration_number).first()
            
            if policy_document:  #First check if found
                policy_start_date_str = policy_document.policy_start_date
                policy_start_date_obj = datetime.strptime(policy_start_date_str, '%Y-%m-%d %H:%M:%S') if policy_start_date_str else None

                return JsonResponse({
                    'success': True,
                    'insurance_company': policy_document.insurance_provider,
                    'policy_number': policy_document.policy_number,
                    'policy_type': policy_document.policy_type,
                    'vehicle_type': policy_document.vehicle_type,
                    'sum_insured': policy_document.sum_insured,
                    'policy_date': policy_start_date_obj.strftime('%Y-%m-%d %H:%M:%S') if policy_start_date_obj else "",
                })
            else:
                # ✅ If no policy found
                return JsonResponse({'success': False, 'message': 'Please fill in all details'})
                
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request'})
#state
def get_state(request):
    states = State.objects.all()
    return render(request, 'leads/create-location-info.html', {'states': states})

def get_cities(request):
    """state_name = request.GET.get('state_id')
    try:
        state = State.objects.get(name=state_name)
        cities = City.objects.filter(state=state).values('city')
        return JsonResponse(list(cities), safe=False)
    except State.DoesNotExist:
        return JsonResponse([], safe=False)"""
    state_id = request.GET.get('state_id')
    cities = City.objects.filter(state_id=state_id).values('id', 'city')
    return JsonResponse({'cities': list(cities)})

@login_required
def bulk_upload_leads(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        camp_name = request.POST.get("camp_name")

        if not excel_file.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, "Only Excel files (.xlsx, .xls) are allowed!")
            return redirect('bulk-upload-leads')

        if not camp_name:
            messages.error(request, "Campaign Name is mandatory.")
            return redirect('bulk-upload-leads')


        # Save the file record
        instance = LeadUploadExcel.objects.create(
            file=excel_file,
            file_name=excel_file.name,
            file_url=excel_file.name,
            campaign_name=camp_name,
            created_by=request.user  # assuming request.user is correct
        )

        # Trigger background task
        async_task("empPortal.taskz.process_lead_excel.process_lead_excel", instance.id)

        messages.success(request, "File uploaded. It will be processed in the background.")
        return redirect("leads-mgt")

    return render(request, "leads/bulk_upload.html")

#for ctraye lead step by step
def lead_init_view(request):
    if not request.user.is_authenticated and request.user.is_active!=1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    types = InsuranceType.objects.all()
    return render(request, 'leads/lead-init.html', {'types': types})

def lead_init_edit(request,lead_id):
    if not request.user.is_authenticated and request.user.is_active!=1: 
        messages.error(request,'Please Login First')
        return redirect('login')
    
    types = InsuranceType.objects.all()
    
    lead_data =  Leads.objects.filter(lead_id=lead_id).first()
    if not lead_data:
        messages.error(request,'This lead is not found in our data')
        return redirect('leads-mgt')
    
    return render(request, 'leads/edit-lead-init.html', {'types': types,'lead_data':lead_data})

def basic_info(request,lead_id):
    if not request.user.is_authenticated and request.user.is_active!=1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    if not lead_id:
        messages.error(request,'Sorry Lead Id is missing')
        return redirect('leads-mgt')
    
    lead_data = Leads.objects.filter(lead_id=lead_id).first()
    if not lead_data:
        messages.error(request,'Sorry Lead Data is missing')
        return redirect('leads-mgt')
    
    customer_data = Customer.objects.filter(id=lead_data.lead_customer_id).last()
    return render(request, "leads/create-basic-details.html", {'lead_data': lead_data,'customer_data':customer_data})
    
def lead_source(request,lead_id):
    if not request.user.is_authenticated and request.user.is_active!=1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    if not lead_id:
        messages.error(request,'Sorry Lead Id is missing')
        return redirect('leads-mgt')
    
    lead_data = Leads.objects.filter(lead_id=lead_id).first()
    if not lead_data:
        messages.error(request,'Sorry Lead Data is missing')
        return redirect('leads-mgt')
    
    source_list = SourceMaster.objects.filter(status=True)
    referral_list = Referral.objects.filter(active=True)
    
    return render(request, "leads/create-lead-source-info.html",{"lead_data":lead_data,"source_list":source_list,"referral_list":referral_list})

def lead_location(request,lead_id):
    if not request.user.is_authenticated and request.user.is_active!=1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    if not lead_id:
        messages.error(request,'Sorry Lead Id is missing')
        return redirect('leads-mgt')
    
    lead_data = Leads.objects.filter(lead_id=lead_id).first()
    if not lead_data:
        messages.error(request,'Sorry Lead Data is missing')
        return redirect('leads-mgt')
    
    states = State.objects.all()
    
    return render(request, "leads/create-location-info.html",{"lead_data":lead_data,"states":states})

def lead_assignment(request,lead_id):
    if not request.user.is_authenticated and request.user.is_active!=1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    if not lead_id:
        messages.error(request,'Sorry Lead Id is missing')
        return redirect('leads-mgt')
    
    lead_data = Leads.objects.filter(lead_id=lead_id).first()
    if not lead_data:
        messages.error(request,'Sorry Lead Data is missing')
        return redirect('leads-mgt')

    user_role = request.user.role_id
    user_dept = request.user.department_id
    
    if user_role == 1:
        assigner_list = Users.objects.filter(is_active=1)
    else:
        assigner_list = Users.objects.filter(is_active=1,department_id=user_dept)   
        
    branches = Branch.objects.filter(status='Active') 
    return render(request, "leads/create-assignment.html",{"lead_data":lead_data,"branches":branches,"assigner_list":assigner_list})

def lead_allocation(request,lead_id):
    if not request.user.is_authenticated and request.user.is_active!=1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    if not lead_id:
        messages.error(request,'Sorry Lead Id is missing')
        return redirect('leads-mgt')
    
    lead_data = Leads.objects.filter(lead_id=lead_id).first()
    if not lead_data:
        messages.error(request,'Sorry Lead Data is missing')
        return redirect('leads-mgt')

    user_role = request.user.role_id
    user_dept = request.user.department_id
    
    if user_role == 1:
        assigner_list = Users.objects.filter(is_active=1)
    else:
        assigner_list = Users.objects.filter(is_active=1,department_id=user_dept)   
        
    branches = Branch.objects.filter(status='Active') 
    return render(request, "leads/lead_allocation.html",{"lead_data":lead_data,"branches":branches,"assigner_list":assigner_list})

def previous_policy_info(request,lead_id):
    if not request.user.is_authenticated and request.user.is_active!=1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    if not lead_id:
        messages.error(request,'Sorry Lead Id is missing')
        return redirect('leads-mgt')
    
    lead_data = Leads.objects.filter(lead_id=lead_id).first()
    if not lead_data:
        messages.error(request,'Sorry Lead Data is missing')
        return redirect('leads-mgt')
    
    insurance_company_list = Insurance.objects.filter(active='Active')
    policy_type_list = PolicyType.objects.filter(status=1)
    vehicle_type_list = VehicleType.objects.filter(status=1)
    
    lead_previous_policy = None
    if lead_data.lead_insurance_product_id == 32:
        lead_previous_policy = LeadPreviousPolicy.objects.filter(lead_id=lead_data.id).last()
    return render(request, "leads/create.html",{"lead_data":lead_data,"lead_previous_policy":lead_previous_policy,'insurance_company_list':insurance_company_list,'policy_type_list':policy_type_list,'vehicle_type_list':vehicle_type_list})

def clean(val):
    return val.strip() if isinstance(val, str) and val.strip() else None

def save_leads_insurance_info(request):
    if not request.user.is_authenticated and request.user.is_active != 1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    lead_insurance_type_id = request.POST.get('insurance_type') or None
    lead_insurance_category_id = request.POST.get('insurance_category') or None
    lead_insurance_product_id = request.POST.get('insurance_product') or None
    lead_first_name = request.POST.get('first_name') or None
    lead_last_name = request.POST.get('last_name') or None
    mobile_number = request.POST.get('mobile') or None
    
    
    customer_data = {
        'customer_f_name': lead_first_name,
        'customer_l_name': lead_last_name,
        'mobile_number': mobile_number,
    }

    customer_id = customer_details(customer_data)
    
    # Apply cleaning
    lead_insurance_type_id = clean(lead_insurance_type_id)
    lead_insurance_category_id = clean(lead_insurance_category_id)
    lead_insurance_product_id = clean(lead_insurance_product_id)
    lead_first_name = clean(lead_first_name)
    lead_last_name = clean(lead_last_name)
    mobile_number = clean(mobile_number)
    
    try:
        parent_existing_lead = Leads.objects.filter(lead_customer_id=customer_id,lead_insurance_product_id=lead_insurance_product_id,parent_lead_id__isnull=True).last()
        if parent_existing_lead:
            if parent_existing_lead.status == True:
                parent_existing_lead.status = False
                parent_existing_lead.save()
            
        existing_lead = Leads.objects.filter(lead_customer_id=customer_id,lead_insurance_product_id=lead_insurance_product_id).last()
        if existing_lead:
            existing_lead.status = False
            existing_lead.save()
            leads_insert = clone_lead(existing_lead, request.user.id,parent_existing_lead)
            lead_ref_id = leads_insert.lead_id
            
            message = f"New lead is created and reborn from lead: #{existing_lead.lead_id}"
        else:
            leads_insert = Leads.objects.create(
                lead_id = int(time.time()),
                lead_insurance_type_id = lead_insurance_type_id,
                lead_insurance_category_id = lead_insurance_category_id,
                lead_insurance_product_id = lead_insurance_product_id,
                lead_customer_id = customer_id,
                lead_first_name = lead_first_name,
                lead_last_name = lead_last_name,
                name_as_per_pan = lead_first_name +' '+ lead_last_name,
                mobile_number = mobile_number,
                created_by_id = request.user.id
            )
            
            lead_ref_id = leads_insert.lead_id
            
            message = 'New lead is created'
        
        LeadActivity.objects.create(
            lead_id = leads_insert.id,
            lead_ref_id = leads_insert.lead_id,
            message = message,
            created_by_id = request.user.id,
        )
        
        messages.success(request,f"Saved Succesfully")
        return redirect('basic-info',lead_id=lead_ref_id)
        
    except Exception as e:
        logger.error(f"Error in save_leads_insurance_info error: {str(e)}")
        messages.error(request,f'Something Went Wrong Please Try After Sometime {str(e)}')
        return redirect('leads-mgt')
    
def clone_lead(existing_lead, created_by_id, parent_existing_lead):
    return Leads.objects.create(
        lead_id = int(time.time()),
        lead_insurance_type_id = existing_lead.lead_insurance_type_id,
        lead_insurance_category_id = existing_lead.lead_insurance_category_id,
        lead_insurance_product_id = existing_lead.lead_insurance_product_id,
        lead_customer_id = existing_lead.lead_customer_id,
        lead_first_name = existing_lead.lead_first_name,
        lead_last_name = existing_lead.lead_last_name,
        name_as_per_pan = existing_lead.name_as_per_pan,
        mobile_number = existing_lead.mobile_number,
        created_by_id = created_by_id,
        parent_lead_id = parent_existing_lead.id,
        
        email_address =  existing_lead.email_address,
        lead_customer_gender =  existing_lead.lead_customer_gender,
        date_of_birth =  existing_lead.date_of_birth,
        lead_customer_identity_no =  existing_lead.lead_customer_identity_no,
        lead_source_type_id = existing_lead.lead_source_type_id,           
        lead_source = existing_lead.lead_source,           
        referral_by = existing_lead.referral_by,           
        referral_name = existing_lead.referral_name,           
        referral_mobile_no = existing_lead.referral_mobile_no,           
        posp_id = existing_lead.posp_id,
        lead_source_medium = existing_lead.lead_source_medium, 
        state_id = existing_lead.state_id,
        city_id = existing_lead.city_id,
        pincode = existing_lead.pincode,
        previous_insurer_name = existing_lead.previous_insurer_name,
        policy_number = existing_lead.policy_number,
        policy_type = existing_lead.policy_type,
        policy_date = existing_lead.policy_date,
        policy_end_date = existing_lead.policy_end_date,
        expiry_status = existing_lead.expiry_status,
        ncb = existing_lead.ncb,
        previous_idv_amount = existing_lead.previous_idv_amount,
        previous_sum_insured = existing_lead.previous_sum_insured,
        claim_made = existing_lead.claim_made,
        claim_amount = existing_lead.claim_amount,
        previous_policy_source = existing_lead.previous_policy_source,
        vehicle_type = existing_lead.vehicle_type,
        vehicle_class = existing_lead.vehicle_class,
        insurance_type = existing_lead.insurance_type,
        product_category = existing_lead.product_category,
        vehicle_reg_no = existing_lead.vehicle_reg_no,
        vehicle_make = existing_lead.vehicle_make,
        vehicle_model = existing_lead.vehicle_model,
        mgf_year = existing_lead.mgf_year,
        sum_insured = existing_lead.sum_insured,          
    )
    
def update_leads_insurance_info(request):
    if not request.user.is_authenticated and request.user.is_active != 1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    lead_id = request.POST.get('lead_ref_id') or None
    lead_insurance_type_id = request.POST.get('insurance_type') or None
    lead_insurance_category_id = request.POST.get('insurance_category') or None
    lead_insurance_product_id = request.POST.get('insurance_product') or None
    lead_first_name = request.POST.get('first_name') or None
    lead_last_name = request.POST.get('last_name') or None
    mobile_number = request.POST.get('mobile') or None
    
    # Apply cleaning
    lead_insurance_type_id = clean(lead_insurance_type_id)
    lead_insurance_category_id = clean(lead_insurance_category_id)
    lead_insurance_product_id = clean(lead_insurance_product_id)
    lead_first_name = clean(lead_first_name)
    lead_last_name = clean(lead_last_name)
    mobile_number = clean(mobile_number)
    
    if not lead_id or lead_id == 0:
        messages.error(request,'Lead Id is not found') 
        return redirect('leads-mgt')
    
    lead_data = Leads.objects.filter(lead_id = lead_id).first()
    
    # try:
    lead_data.lead_insurance_type_id = int(lead_insurance_type_id)
    lead_data.lead_insurance_category_id = int(lead_insurance_category_id)
    lead_data.lead_insurance_product_id = int(lead_insurance_product_id)
    lead_data.lead_first_name = lead_first_name
    lead_data.lead_last_name = lead_last_name
    lead_data.mobile_number = mobile_number
    lead_data.save()
    
    lead_id = lead_data.lead_id
    messages.success(request,f"Saved Succesfully")
    return redirect('basic-info',lead_id=lead_id)
        
    # except Exception as e:
    #     logger.error(f"Error in update_leads_insurance_info error: {str(e)}")
    #     messages.error(request,'Something Went Wrong Please Try After Sometime')
    #     return redirect('leads-mgt')
    
def save_leads_basic_info(request):
    if not request.user.is_authenticated and request.user.is_active != 1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    lead_id = request.POST.get('lead_id') or None
    customer_name = request.POST.get('customer_name') or None
    email_address = request.POST.get('email_address') or None
    gender = request.POST.get('gender') or None
    date_of_birth = request.POST.get('date_of_birth') or None
    identity_no = request.POST.get('identity_no') or None
    
    if not lead_id or lead_id == 0:
        messages.error(request,'Lead Id is not found') 
        return redirect('leads-mgt')
    
    lead_data = Leads.objects.filter(lead_id = lead_id).first()
    try:
        lead_customer_id = lead_data.lead_customer_id
        if lead_customer_id:
            customer_data = Customer.objects.filter(id = lead_customer_id).last()
            customer_data.email_address = clean(email_address)
            customer_data.date_of_birth = clean(date_of_birth)
            customer_data.gender = clean(gender)
            customer_data.identity_no = clean(identity_no)
            customer_data.save()
        
        lead_data.name_as_per_pan = clean(customer_name)
        lead_data.email_address = clean(email_address)
        lead_data.lead_customer_gender = clean(gender)
        lead_data.date_of_birth = clean(date_of_birth)
        lead_data.lead_customer_identity_no = clean(identity_no)
        lead_data.save()
        
        lead_id = lead_data.lead_id
        messages.success(request,f"Saved Succesfully")
        return redirect('lead-source',lead_id=lead_id)
        
    except Exception as e:
        logger.error(f"Error in save_leads_basic_info error: {str(e)}")
        messages.error(request,'Something Went Wrong Please Try After Sometime')
        return redirect('leads-mgt')
    
def save_leads_source_info(request):
    if not request.user.is_authenticated and request.user.is_active != 1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    lead_id = request.POST.get('lead_ref_id') or None
    lead_source_type = request.POST.get('lead_source_type') or None
    lead_source = request.POST.get('lead_source_name') or None
    refered_by = request.POST.get('refered_by') or None
    referral_name = request.POST.get('referral_name') or None
    referral_mobile_number = request.POST.get('referral_mobile_number') or None
    posp_by = request.POST.get('posp_by') or None
    source_medium = request.POST.get('source_medium') or None
    
    if not lead_id or lead_id == 0:
        messages.error(request,'Lead Id is not found') 
        return redirect('leads-mgt')
    
    lead_data = Leads.objects.filter(lead_id = lead_id).first()
    try:
        lead_data.lead_source_type_id = clean(lead_source_type)
        lead_data.lead_source = clean(lead_source)
        lead_data.referral_by = clean(refered_by)
        lead_data.referral_name = clean(referral_name)
        lead_data.referral_mobile_no = clean(referral_mobile_number)
        lead_data.posp_id = int(posp_by) if posp_by else None
        lead_data.lead_source_medium = clean(source_medium)
        lead_data.save()
        
        lead_ref_id = lead_data.lead_id
        messages.success(request,f"Saved Succesfully")
        return redirect('lead-location',lead_id=lead_ref_id)
        
    except Exception as e:
        logger.error(f"Error in save_leads_source_info error: {str(e)}")
        messages.error(request,'Something Went Wrong Please Try After Sometime')
        return redirect('leads-mgt')
    
def save_leads_location_info(request):
    if not request.user.is_authenticated and request.user.is_active != 1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    lead_id = request.POST.get('lead_ref_id') or None
    state = request.POST.get('state') or None
    city = request.POST.get('city') or None
    pincode = request.POST.get('pincode') or None
    
    if not lead_id or lead_id == 0:
        messages.error(request,'Lead Id is not found') 
        return redirect('leads-mgt')
    
    lead_data = Leads.objects.filter(lead_id = lead_id).first()
    try:
        lead_data.state_id = clean(state)
        lead_data.city_id = clean(city)
        lead_data.pincode = clean(pincode)
        lead_data.save()
        
        lead_ref_id = lead_data.lead_id
        messages.success(request,f"Saved Succesfully")
        return redirect('lead-assignment',lead_id=lead_ref_id)
        
    except Exception as e:
        logger.error(f"Error in save_leads_location_info error: {str(e)}")
        messages.error(request,'Something Went Wrong Please Try After Sometime')
        return redirect('leads-mgt')
    
def save_leads_assignment_info(request):
    if not request.user.is_authenticated and request.user.is_active != 1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    lead_id = request.POST.get('lead_ref_id') or None
    assigned_to = request.POST.get('assigned_to') or None
    assigned_manager = request.POST.get('assigned_manager') or None
    assigned_teamleader = request.POST.get('assigned_teamleader') or None
    branch = request.POST.get('branch') or None
    lead_status_type = request.POST.get('lead_status_type') or None
    lead_tag = request.POST.get('lead_tag') or None
    
    if not lead_id or lead_id == 0:
        messages.error(request,'Lead Id is not found') 
        return redirect('leads-mgt')
    
    lead_data = Leads.objects.filter(lead_id = lead_id).first()
    try:
        lead_data.assigned_to_id = assigned_to
        lead_data.assigned_manager_id = assigned_manager
        lead_data.assigned_teamleader_id = assigned_teamleader
        lead_data.branch_id = clean(branch)
        lead_data.lead_status_type = clean(lead_status_type)
        lead_data.lead_tag = clean(lead_tag)
        lead_data.save()
        
        lead_ref_id = lead_data.lead_id
        
        message = f'Lead is allocated to {lead_data.assigned_to.full_name} in branch {lead_data.branch.branch_name}'
        LeadActivity.objects.create(
            lead_id = lead_data.id,
            lead_ref_id = lead_data.lead_id,
            message = message,
            created_by_id = request.user.id,
        )
         
        messages.success(request,f"Saved Succesfully")
        return redirect('leads-previous-policy-info',lead_id=lead_ref_id)
        
    except Exception as e:
        logger.error(f"Error in save_leads_assignment_info error: {str(e)}")
        messages.error(request,'Something Went Wrong Please Try After Sometime')
        return redirect('leads-mgt')
       
def save_leads_allocation_info(request):
    if not request.user.is_authenticated and request.user.is_active != 1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    lead_id = request.POST.get('lead_ref_id') or None
    assigned_to = request.POST.get('assigned_to') or None
    assigned_manager = request.POST.get('assigned_manager') or None
    assigned_teamleader = request.POST.get('assigned_teamleader') or None
    branch = request.POST.get('branch') or None
    lead_status_type = request.POST.get('lead_status_type') or None
    lead_tag = request.POST.get('lead_tag') or None
    
    if not lead_id or lead_id == 0:
        messages.error(request,'Lead Id is not found') 
        return redirect('leads-mgt')
    
    lead_data = Leads.objects.filter(lead_id = lead_id).first()
    try:
        lead_data.assigned_to_id = assigned_to
        lead_data.assigned_manager_id = assigned_manager
        lead_data.assigned_teamleader_id = assigned_teamleader
        lead_data.branch_id = clean(branch)
        lead_data.lead_status_type = clean(lead_status_type)
        lead_data.lead_tag = clean(lead_tag)
        lead_data.save()
        
        lead_ref_id = lead_data.lead_id
        
        message = f'Lead is allocated to {lead_data.assigned_to.full_name} in branch {lead_data.branch.branch_name}'
        LeadActivity.objects.create(
            lead_id = lead_data.id,
            lead_ref_id = lead_ref_id,
            message = message,
            created_by_id = request.user.id,
        )
         
        messages.success(request,f"Allocated Succesfully")
        return redirect('leads-mgt')
        
    except Exception as e:
        logger.error(f"Error in save_leads_allocation_info error: {str(e)}")
        messages.error(request,'Something Went Wrong Please Try After Sometime')
        return redirect('leads-mgt')
    
def save_leads_previous_policy_info(request):
    if not request.user.is_authenticated and request.user.is_active != 1:
        messages.error(request,'Please Login First')
        return redirect('login')
    
    lead_id = request.POST.get('lead_ref_id') or None
    previous_insurer_name = request.POST.get('previous_insurer_name') or None
    policy_number = request.POST.get('policy_number') or None
    policy_type = request.POST.get('policy_type') or None
    policy_date = request.POST.get('policy_date') or None
    policy_end_date = request.POST.get('policy_end_date') or None
    expiry_status = request.POST.get('expiry_status') or None
    ncb = request.POST.get('ncb') or None
    previous_idv_amount = request.POST.get('previous_idv_amount') or None
    previous_sum_insured = request.POST.get('previous_sum_insured') or None
    claim_made = request.POST.get('claim_made') or None
    claim_amount = request.POST.get('claim_amount') or None
    previous_policy_source = request.POST.get('previous_policy_source') or None
    vehicle_type = request.POST.get('vehicle_type') or None
    vehicle_class = request.POST.get('vehicle_class') or None
    insurance_type = request.POST.get('insurance_type') or None
    product_category = request.POST.get('product_category') or None
    vehicle_reg_no = request.POST.get('vehicle_reg_no') or None
    vehicle_make = request.POST.get('vehicle_make') or None
    vehicle_model = request.POST.get('vehicle_model') or None
    mgf_year = request.POST.get('mgf_year') or None
    sum_insured = request.POST.get('sum_insured') or None
    
    if not lead_id or lead_id == 0:
        messages.error(request,'Lead Id is not found') 
        return redirect('leads-mgt')
    
    lead_data = Leads.objects.filter(lead_id = lead_id).first()
    try:
        lead_data.insurance_company = clean(previous_insurer_name)
        lead_data.policy_number = clean(policy_number)
        lead_data.policy_type = clean(policy_type)
        lead_data.policy_date = clean(policy_date)
        lead_data.policy_end_date = clean(policy_end_date)
        lead_data.expiry_status = clean(expiry_status)
        lead_data.ncb = clean(ncb)
        lead_data.previous_idv_amount = clean(previous_idv_amount)
        lead_data.previous_sum_insured = clean(previous_sum_insured)
        lead_data.claim_made = clean(claim_made)
        lead_data.claim_amount = clean(claim_amount)
        lead_data.previous_policy_source = clean(previous_policy_source)
        lead_data.vehicle_type = clean(vehicle_type)
        lead_data.vehicle_class = clean(vehicle_class)
        lead_data.insurance_type = clean(insurance_type)
        lead_data.product_category = clean(product_category)
        lead_data.vehicle_reg_no = clean(vehicle_reg_no)
        lead_data.vehicle_make = clean(vehicle_make)
        lead_data.vehicle_model = clean(vehicle_model)
        lead_data.mgf_year = clean(mgf_year)
        lead_data.sum_insured = clean(sum_insured)
        lead_data.save()
            
        messages.success(request,f"Saved Succesfully")
        return redirect('leads-mgt')
        
    except Exception as e:
        logger.error(f"Error in save_leads_previous_policy_info error: {str(e)}")
        messages.error(request,'Something Went Wrong Please Try After Sometime')
        return redirect('leads-mgt')

def save_leads_motor_previous_policy_info(request):
    if not request.user.is_authenticated or request.user.is_active != 1:
        messages.error(request, 'Please Login First')
        return redirect('login')

    lead_ref_id = request.POST.get('lead_ref_id') or None
    if not lead_ref_id or lead_ref_id == "0":
        messages.error(request, 'Lead ID is not found')
        return redirect('leads-mgt')

    lead = Leads.objects.filter(lead_id=lead_ref_id).first()
    if not lead:
        messages.error(request, 'Invalid Lead Reference')
        return redirect('leads-mgt')

    policy, created = LeadPreviousPolicy.objects.get_or_create(lead_id=lead.id)

    fields = [
        'registration_number', 'registration_date', 'make', 'model', 'variant',
        'year_of_manufacture', 'registration_state', 'registration_city', 'chassis_number',
        'engine_number', 'claim_history', 'ncb', 'ncb_percentage', 'idv_value',
        'policy_duration', 'addons', 'owner_name', 'father_name', 'state_code', 'location',
        'vehicle_category', 'vehicle_class_description', 'body_type_description', 'vehicle_color',
        'vehicle_cubic_capacity', 'vehicle_gross_weight', 'vehicle_seating_capacity',
        'vehicle_fuel_description', 'vehicle_owner_number', 'rc_expiry_date',
        'rc_pucc_expiry_date', 'insurance_expiry_date', 'insurance_policy_number'
    ]

    for field in fields:
        value = request.POST.get(field) or None
        setattr(policy, field, clean(value))

    try:
        insurance_company = request.POST.get('insurance_company')
        vehicle_type = request.POST.get('vehicle_type')
        policy_type = request.POST.get('policy_type')
        policy.insurance_company_id = clean(insurance_company)
        policy.vehicle_type_id = clean(vehicle_type)
        policy.policy_type_id = clean(policy_type)
        policy.save()
        messages.success(request, f"{'Created' if created else 'Updated'} successfully.")
        return redirect('leads-mgt')
    except Exception as e:
        logger.error(f"Error saving LeadPreviousPolicy: {str(e)}")
        messages.error(request, 'Something went wrong. Please try again later.')
        return redirect('leads-mgt')

def view_lead(request, lead_id):
    if not request.user.is_authenticated and request.user.is_active != 1:
        return redirect('login')
        
    lead = get_object_or_404(Leads, lead_id=lead_id)
    
    disposition_list = Disposition.objects.filter(disp_is_active=True)
    lead_disposition = LeadDisposition.objects.filter(lead_id=lead.id).last()
    disposition_logs  = LeadDispositionLogs.objects.filter(log_lead_id=lead.id).order_by('-log_created_at')
    lead_previous_policy = None
    if lead.lead_insurance_product_id == 32:
        lead_previous_policy = LeadPreviousPolicy.objects.filter(lead_id=lead.id).last()
    return render(request, 'leads/lead-view.html', {'lead': lead,'disposition_list':disposition_list,'disposition_logs':disposition_logs,'lead_disposition':lead_disposition,'lead_previous_policy':lead_previous_policy})

def save_leads_dispositions(request):
    if not request.user.is_authenticated and request.user.is_active != 1:
        messages.error(request,'Please Login First')
        return JsonResponse({'statusCode': 401, 'status': 'Authentication Failed', 'message': 'Please Login First'})
        
    if request.method == "POST":
        lead_ref_id = request.POST.get('lead_id')
        main_disposition = request.POST.get('main_disposition')
        sub_disposition = request.POST.get('sub_disposition')
        
        follow_up_date = request.POST.get('follow_up_date',None)
        if follow_up_date:
            try:
                follow_up_date = datetime.strptime(follow_up_date, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, 'Invalid date format. Please use YYYY-MM-DD.')
                return JsonResponse({
                    'statusCode': 400,
                    'status': 'failed',
                    'message': 'Invalid date format. It must be in YYYY-MM-DD format.'
                })
        else:
            follow_up_date = None
            
        follow_up_time = request.POST.get('follow_up_time', None)
        if follow_up_time:
            try:
                follow_up_time = datetime.strptime(follow_up_time, '%H:%M').time()
            except ValueError:
                try:
                    follow_up_time = datetime.strptime(follow_up_time, '%H:%M:%S').time()
                except ValueError:
                    messages.error(request, 'Invalid time format. Please use HH:MM or HH:MM:SS.')
                    return JsonResponse({
                        'statusCode': 400,
                        'status': 'failed',
                        'message': 'Invalid time format. It must be in HH:MM[:ss] format.'
                    })
        else:
            # Handle None case if follow_up_time is not provided
            follow_up_time = None
            
        remark = request.POST.get('remark',None)

        if not (lead_ref_id and main_disposition and sub_disposition):
            messages.error(request, 'All fields are required.')
            return JsonResponse({'statusCode': 405, 'status': 'failed', 'message': 'Missing required fields'})

        lead_data = Leads.objects.filter(lead_id=lead_ref_id).last()
        lead_id = lead_data.id
        if not lead_data:
            messages.error(request, 'Lead id is not found')
            return JsonResponse({'statusCode': 2, 'status': 'failed', 'message': 'Missing lead id'})

        try:
            lead_disp_data = LeadDisposition.objects.filter(lead_id=lead_id).first()
            if lead_disp_data:
                dispo_name = lead_disp_data.disp.disp_name
                sub_disp_name = lead_disp_data.sub_disp.sub_disp_name
                
                lead_disp_data.disp_id = main_disposition
                lead_disp_data.sub_disp_id = sub_disposition
                lead_disp_data.updated_by_id = request.user.id
                lead_disp_data.followup_date = follow_up_date
                lead_disp_data.followup_time = follow_up_time
                lead_disp_data.remark = remark
                lead_disp_data.save()
                status = "Updated"
                message = f'Lead is Disposed by {lead_disp_data.created_by.full_name} , disposition form {dispo_name} to {lead_disp_data.disp.disp_name} and  sub disposition from {sub_disp_name} to {lead_disp_data.sub_disp.sub_disp_name} '
                
            else:
                lead_disp_data = LeadDisposition.objects.create(
                    lead_id = lead_id,
                    disp_id = main_disposition, 
                    sub_disp_id = sub_disposition, 
                    created_by_id = request.user.id, 
                    followup_date = follow_up_date,
                    followup_time = follow_up_time,
                    remark = remark
                )
                status = "Created"
                message = f'Lead is Disposed by {lead_disp_data.created_by.full_name} , disposition name: {lead_disp_data.disp.disp_name}, sub disposition name: {lead_disp_data.sub_disp.sub_disp_name} '
                
                 
            logs = LeadDispositionLogs.objects.create(
                log_lead_disp_id = lead_disp_data.id,
                log_lead_id  = lead_id,
                log_disp_id  = main_disposition,
                log_sub_disp_id  = sub_disposition,
                log_created_by_id  = request.user.id,
                log_followup_date = follow_up_date,
                log_followup_time = follow_up_time,
                log_remark = remark
            )
            
            LeadActivity.objects.create(
                lead_id = lead_data.id,
                lead_ref_id = lead_data.lead_id,
                message = message,
                created_by_id = request.user.id,
            )
        
            messages.success(request,'Saved Successfully')
            return JsonResponse({
                    'statusCode': 200,
                    'status': 'success',
                    'action': status,
                    'lead_disposition_id': lead_disp_data.id
                })
        except Exception as e:
            logger.error(f"Failed to insert or update disposition for lead_id {lead_id}. Error: {str(e)}")
            messages.error(request, 'Something went wrong')
            return JsonResponse({
                'statusCode': 500,
                'status': 'failed',
                'message': str(e)
            })
    else:
        logger.error(f"Failed to insert or update disposition for lead_id Error")
        messages.error(request,'Something went Wrong')
        return JsonResponse({
            'statusCode': 405,
            'status': 'failed',
            'message': 'Invalid request method'
        })
    
def customer_details(data):
    mobile = data.get('mobile_number')

    if not mobile:
        return None
    
    customer = Customer.objects.filter(mobile_number=mobile).first()

    if customer:
        return customer.id

    customer = Customer.objects.create(
        name_as_per_pan=f"{data.get('customer_f_name', '')} {data.get('customer_l_name', '')}".strip(),
        mobile_number=mobile
    )

    return customer.id

def get_lead_activity_logs(request):
    if request.user.is_authenticated and request.user.is_active == 1:
        lead_id = request.POST.get('lead_id')
        activity_logs = LeadActivity.objects.filter(lead_ref_id=lead_id).order_by('-created_at')

        html = ''
        if activity_logs.exists():
            for activity in activity_logs:
                html += f'''
                    <tr>
                        <td>
                            <div class="d-flex">
                                <div class="activity-circle">
                                    {activity.created_by.full_name[0].upper() if activity.created_by and activity.created_by.full_name else 'U'}
                                </div>
                                <div class="ml-2">
                                    <span class="d-block">{activity.create_date}</span>
                                    <span class="text-muted">{activity.create_time}</span>
                                </div>    
                            </div>    
                        </td>
                        <td>
                            <div class="d-flex flex-wrap">
                                <span class="d-block text-primary">{activity.message}</span>
                            </div>    
                            <span class="text-muted d-block">
                                Added by {activity.created_by.full_name if activity.created_by else 'Unknown'} on {activity.created_at.strftime('%d %b %Y %I:%M %p')}
                            </span>
                        </td>
                    </tr>
                '''
        else:
            html = '''
                <tr>
                    <td colspan="2" class="text-center text-muted">No data found</td>
                </tr>
            '''
        return JsonResponse({'html': html})
    else:
        return JsonResponse({'html': 'Please Login First'})
    