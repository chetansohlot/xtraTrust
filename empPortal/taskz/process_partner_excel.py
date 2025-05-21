from ..models import UploadedZip, FileAnalysis, ExtractedFile, BulkPolicyLog, Commission, PolicyDocument, UnprocessedPolicyFiles, ChatGPTLog, UploadedExcel
from ..models import PolicyInfo, PolicyVehicleInfo, AgentPaymentDetails, InsurerPaymentDetails, FranchisePayment
from ..models import FranchisePaymentLog, PolicyInfoLog, PolicyVehicleInfoLog, AgentPaymentDetailsLog, InsurerPaymentDetailsLog
import django, dramatiq, fitz, os, zipfile, requests, re, json, traceback, time, logging, shutil
from django.conf import settings
from django.utils import timezone
from django.utils.timezone import now
from django_q.tasks import async_task
OPENAI_API_KEY = settings.OPENAI_API_KEY
from django.db.models import F
from ..utils import getUserNameByUserId, commisionRateByMemberId, insurercommisionRateByMemberId, to_int
from django.core.exceptions import ObjectDoesNotExist
from django.utils.encoding import filepath_to_uri
import pandas as pd
import openpyxl, re, datetime, logging
from django.utils import timezone
from dateutil import parser
from django.contrib.auth.hashers import make_password
from ..models import PartnerUploadExcel, Users

logger = logging.getLogger(__name__)

def process_partner_excel(file_id):
    try:
        instance = PartnerUploadExcel.objects.get(id=file_id)
        wb = openpyxl.load_workbook(instance.file.path)
        sheet = wb.active

        total_rows = sheet.max_row - 1
        success, error = 0, 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                full_name, gender_str, email, mobile, password, dob, pan_no = [str(cell).strip() if cell else '' for cell in row]

                if not full_name:
                    error += 1
                    continue

                name_parts = full_name.split()
                first_name = name_parts[0]
                last_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else ""

                if gender_str.lower() not in ['male', 'female']:
                    error += 1
                    continue
                gender = 1 if gender_str.lower() == "male" else 2

                if not email or not re.match(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$", email):
                    error += 1
                    continue

                if not mobile or not mobile.isdigit() or len(mobile) != 10 or mobile[0] not in "6789":
                    error += 1
                    logger.warning(f"Skipped: Invalid mobile - {mobile}")
                    continue

                if not password or len(password) < 6:
                    error += 1
                    continue

                try:
                    dob_date = parser.parse(dob).date()
                    today = datetime.date.today()
                    age = (today - dob_date).days // 365
                    if dob_date >= today or age < 18:
                        error += 1
                        continue
                except:
                    error += 1
                    continue

                if not pan_no or not re.match(r'^[A-Z]{5}[0-9]{4}[A-Z]$', pan_no.upper()):
                    error += 1
                    continue

                if Users.objects.filter(email=email).exists() or \
                   Users.objects.filter(phone=mobile).exists() or \
                   Users.objects.filter(pan_no=pan_no.upper()).exists():
                    error += 1
                    continue

                last_user = Users.objects.order_by("-id").first()
                user_gen_id = f"UR-{(int(last_user.user_gen_id.split('-')[1]) + 1):04d}" if last_user and last_user.user_gen_id.startswith("UR-") else "UR-0001"

                Users.objects.create(
                    user_gen_id=user_gen_id,
                    role_id=4,
                    role_name="User",
                    user_name=full_name,
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                    phone=mobile,
                    gender=gender,
                    password=make_password(password),
                    dob=dob_date,
                    pan_no=pan_no.upper(),
                    status=1,
                    is_active=1
                )

                success += 1

            except Exception as e:
                logger.error(f"Error processing row: {e}")
                error += 1

        instance.total_rows = total_rows
        instance.success_rows = success
        instance.error_rows = error
        instance.valid_rows = success
        instance.invalid_rows = error
        instance.is_processed = True
        instance.save()

        logger.info(f"Background task finished - Success: {success}, Errors: {error}")

    except Exception as e:
        logger.error(f"Task Failed: {e}")
        instance.is_processed = True
        instance.error = str(e)
        instance.save()
